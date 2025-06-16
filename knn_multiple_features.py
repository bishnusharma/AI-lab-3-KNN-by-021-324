import math
from collections import Counter
from openpyxl import load_workbook


def load_data(filename):
    """Load data from Excel file using openpyxl"""
    wb = load_workbook(filename)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        age, tumor_size, lymph_nodes, cell_uniformity, diagnosis = row
        data.append({
            'features': [age, tumor_size, lymph_nodes, cell_uniformity],
            'diagnosis': diagnosis
        })
    return data


def normalize_data(data):
    """Normalize features to [0,1] range"""
    # Initialize min and max for each feature
    num_features = len(data[0]['features'])
    min_vals = [float('inf')] * num_features
    max_vals = [float('-inf')] * num_features

    # Find min and max for each feature
    for entry in data:
        for i, val in enumerate(entry['features']):
            if val < min_vals[i]:
                min_vals[i] = val
            if val > max_vals[i]:
                max_vals[i] = val

    # Normalize data
    normalized_data = []
    for entry in data:
        normalized_features = []
        for i, val in enumerate(entry['features']):
            if max_vals[i] == min_vals[i]:  # Avoid division by zero
                normalized = 0.0
            else:
                normalized = (val - min_vals[i]) / (max_vals[i] - min_vals[i])
            normalized_features.append(normalized)

        normalized_data.append({
            'features': normalized_features,
            'diagnosis': entry['diagnosis']
        })

    return normalized_data, (min_vals, max_vals)


def manhattan_distance(a, b):
    """Calculate Manhattan distance between two feature vectors"""
    return sum(abs(x - y) for x, y in zip(a, b))


def knn_predict(train_data, test_point, k=19):
    """Predict class using KNN algorithm"""
    distances = []
    for train_point in train_data:
        dist = manhattan_distance(train_point['features'], test_point)
        distances.append((dist, train_point['diagnosis']))

    # Sort by distance and get k nearest neighbors
    distances.sort(key=lambda x: x[0])
    neighbors = distances[:k]

    # Count votes for each class
    votes = [n[1] for n in neighbors]
    vote_count = Counter(votes)

    # Return the class with most votes
    return vote_count.most_common(1)[0][0]


def evaluate_model(train_data, test_data, k=19):
    """Evaluate model accuracy on test data"""
    correct = 0
    for test_point in test_data:
        prediction = knn_predict(train_data, test_point['features'], k)
        if prediction == test_point['diagnosis']:
            correct += 1
    return correct / len(test_data)


def main():
    # Load and prepare data
    data = load_data("breast_cancer_dataset.xlsx")
    normalized_data, norm_params = normalize_data(data)

    # Split data into train and test (80/20)
    split_idx = int(0.8 * len(normalized_data))
    train_data = normalized_data[:split_idx]
    test_data = normalized_data[split_idx:]

    # Evaluate model
    accuracy = evaluate_model(train_data, test_data, k=19)
    print(f"Model Accuracy: {accuracy:.2%}")

    # Example prediction
    new_point = [45, 2.5, 3, 7]  # Age, Tumor Size, Lymph Nodes, Cell Uniformity
    # Normalize new point using same parameters
    min_vals, max_vals = norm_params
    new_point_normalized = [
        (new_point[i] - min_vals[i]) / (max_vals[i] - min_vals[i])
        for i in range(len(new_point))
    ]
    prediction = knn_predict(normalized_data, new_point_normalized, k=19)
    print(f"Prediction for {new_point}: {'Malignant' if prediction == 1 else 'Benign'}")


if __name__ == "__main__":
    main()